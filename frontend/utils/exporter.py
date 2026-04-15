# frontend/utils/exporter.py

import csv
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

class TableExporter:
    @staticmethod
    def export_to_csv(path: str, headers: list, rows: list):
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow(headers)
            writer.writerows(rows)

    @staticmethod
    def export_to_excel(path: str, headers: list, rows: list, sheet_name: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if ws is None: 
            raise Exception("Ошибка создания листа Excel")
            
        ws.title = sheet_name.capitalize()

        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row_data in rows:
            converted_row = [int(v) if str(v).isdigit() else v for v in row_data]
            ws.append(converted_row)

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_length = max((len(str(cell.value)) for cell in col_cells if cell.value), default=0)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        wb.save(path)

    @staticmethod
    def export_to_html(path: str, headers: list, rows: list, title: str):
        html_lines = [
            "<!DOCTYPE html>", "<html><head><meta charset='utf-8'>",
            "<style>table {border-collapse: collapse; width: 100%;} th, td {border: 1px solid black; padding: 8px;} th {background-color: #f2f2f2;}</style>",
            "</head><body>", f"<h2>Отчет: {title.capitalize()}</h2>", "<table>", "<tr>"
        ]
        html_lines.extend(f"<th>{h}</th>" for h in headers)
        html_lines.append("</tr>")
        
        for row in rows:
            html_lines.append("<tr>")
            html_lines.extend(f"<td>{cell}</td>" for cell in row)
            html_lines.append("</tr>")
        html_lines.extend(["</table>", "</body></html>"])
        
        with open(path, 'w', encoding='utf-8') as f:
            f.write("\n".join(html_lines))

    @staticmethod
    def export_to_sql(path: str, headers: list, rows: list, table_name: str):
        cols_str = ", ".join(f'"{h}"' for h in headers)
        sql_lines = []
        for row in rows:
            vals = [v if str(v).isdigit() else f"'{str(v).replace('\'', '\'\'')}'" for v in row]
            sql_lines.append(f"INSERT INTO {table_name} ({cols_str}) VALUES ({', '.join(vals)});")
        
        with open(path, 'w', encoding='utf-8') as f:
            f.write("\n".join(sql_lines))